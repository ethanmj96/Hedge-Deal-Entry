from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from datetime import date
from pathlib import Path
from calendar import monthrange
import openpyxl
import pdfplumber
import re
import io

app = FastAPI(title="Deal Entry API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_FILE = Path(__file__).parent / "data.xlsx"

COLUMNS = [
    "Trade ID", "Trade Date", "Start Date", "End Date", "Counterparty", "Direction",
    "Commodity", "Product", "Trade Type", "Volume",
    "Volume Unit", "Price", "Currency",
]

DROPDOWN_OPTIONS = {
    "counterparties": ["Goldman Sachs", "JP Morgan", "RBC", "Scotiabank", "TD", "ATB"],
    "directions": ["Buy", "Sell"],
    "commodities": ["NA Natural Gas", "European Natural Gas", "Crude", "NGL"],
    "products": ["WTI", "Dated Brent", "AECO 7A", "AECO 5A", "Conway", "TTF", "NBP"],
    "trade_types": ["Swap", "Call", "Put", "Collar", "3-Way"],
    "volume_units": ["GJ", "BBL", "MWh"],
    "currencies": ["CAD", "USD", "EUR", "GBP"],
}

PRODUCT_CURRENCY_MAP = {
    "WTI": "USD",
    "Dated Brent": "USD",
    "AECO 7A": "CAD",
    "AECO 5A": "CAD",
    "Conway": "USD",
    "TTF": "EUR",
    "NBP": None,
}

PRODUCT_COMMODITY_MAP = {
    "WTI": "Crude",
    "Dated Brent": "Crude",
    "AECO 7A": "NA Natural Gas",
    "AECO 5A": "NA Natural Gas",
    "Conway": "NGL",
    "TTF": "European Natural Gas",
    "NBP": "European Natural Gas",
}

PRODUCT_UNIT_MAP = {
    "WTI": "BBL",
    "Dated Brent": "BBL",
    "AECO 7A": "GJ",
    "AECO 5A": "GJ",
    "Conway": "BBL",
    "TTF": "MWh",
    "NBP": "MWh",
}


class Deal(BaseModel):
    trade_id: str
    trade_date: date
    counterparty: str
    direction: str | None = None
    commodity: str | None = None
    product: str
    trade_type: str
    volume: float
    volume_unit: str | None = None
    price: float | None = None
    currency: str | None = None
    put_strike: float | None = None
    call_strike: float | None = None
    sold_put_strike: float | None = None
    start_date: date
    end_date: date


PRICE_COLUMNS = ["Product", "Month", "Price", "Type", "Date"]


def migrate_sheet(ws, expected_columns):
    """Add any missing columns to an existing sheet without deleting data."""
    existing = [cell.value for cell in ws[1]]
    for col_name in expected_columns:
        if col_name not in existing:
            new_col = len(existing) + 1
            ws.cell(row=1, column=new_col, value=col_name)
            existing.append(col_name)


def get_or_create_workbook():
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deals"
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)
    # Migrate Deals sheet if columns were added
    migrate_sheet(wb["Deals"], COLUMNS)
    # Rename old "Settled Prices" sheet to "Prices" if it exists
    if "Settled Prices" in wb.sheetnames:
        wb["Settled Prices"].title = "Prices"
    if "Prices" not in wb.sheetnames:
        ws2 = wb.create_sheet("Prices")
        ws2.append(PRICE_COLUMNS)
    else:
        migrate_sheet(wb["Prices"], PRICE_COLUMNS)
    wb.save(EXCEL_FILE)
    return wb


@app.get("/")
def health():
    return {"status": "ok"}


@app.get("/options")
def get_options():
    return DROPDOWN_OPTIONS


def make_row(trade_id, deal, direction, trade_type, price):
    return [
        trade_id,
        deal.trade_date.isoformat(),
        deal.start_date.isoformat(),
        deal.end_date.isoformat(),
        deal.counterparty,
        direction,
        deal.commodity,
        deal.product,
        trade_type,
        deal.volume,
        deal.volume_unit,
        price,
        deal.currency,
    ]


@app.post("/deals")
def create_deal(deal: Deal):
    # Auto-fill commodity and currency from product
    if not deal.commodity:
        deal.commodity = PRODUCT_COMMODITY_MAP.get(deal.product)
    if not deal.currency:
        deal.currency = PRODUCT_CURRENCY_MAP.get(deal.product)
    if not deal.volume_unit:
        deal.volume_unit = PRODUCT_UNIT_MAP.get(deal.product)

    wb = get_or_create_workbook()
    ws = wb["Deals"]

    if deal.trade_type == "Collar":
        ws.append(make_row(f"{deal.trade_id}-1", deal, "Buy", "Put", deal.put_strike))
        ws.append(make_row(f"{deal.trade_id}-2", deal, "Sell", "Call", deal.call_strike))
    elif deal.trade_type == "3-Way":
        ws.append(make_row(f"{deal.trade_id}-1", deal, "Buy", "Put", deal.put_strike))
        ws.append(make_row(f"{deal.trade_id}-2", deal, "Sell", "Call", deal.call_strike))
        ws.append(make_row(f"{deal.trade_id}-3", deal, "Sell", "Put", deal.sold_put_strike))
    else:
        ws.append(make_row(deal.trade_id, deal, deal.direction, deal.trade_type, deal.price))

    wb.save(EXCEL_FILE)
    return {"message": "Deal saved", "trade_id": deal.trade_id}


@app.get("/deals")
def list_deals():
    wb = get_or_create_workbook()
    ws = wb["Deals"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    deals = []
    for row in rows:
        if row[0] is None:
            continue
        deals.append(dict(zip(COLUMNS, row)))
    return deals


class DealUpdate(BaseModel):
    trade_date: date
    start_date: date
    end_date: date
    counterparty: str
    direction: str
    commodity: str
    product: str
    trade_type: str
    volume: float
    volume_unit: str
    price: float
    currency: str


@app.put("/deals/{trade_id}")
def update_deal(trade_id: str, update: DealUpdate):
    wb = get_or_create_workbook()
    ws = wb["Deals"]
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == trade_id:
            values = [
                trade_id,
                update.trade_date.isoformat(),
                update.start_date.isoformat(),
                update.end_date.isoformat(),
                update.counterparty,
                update.direction,
                update.commodity,
                update.product,
                update.trade_type,
                update.volume,
                update.volume_unit,
                update.price,
                update.currency,
            ]
            for i, cell in enumerate(row):
                cell.value = values[i]
            wb.save(EXCEL_FILE)
            return {"message": "Deal updated", "trade_id": trade_id}
    raise HTTPException(status_code=404, detail="Deal not found")


@app.delete("/deals/{trade_id}")
def delete_deal(trade_id: str):
    wb = get_or_create_workbook()
    ws = wb["Deals"]
    rows_to_delete = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        val = row[0].value
        if val == trade_id or (val and val.startswith(trade_id + "-")):
            rows_to_delete.append(row_idx)
    if not rows_to_delete:
        raise HTTPException(status_code=404, detail="Deal not found")
    for offset, row_idx in enumerate(rows_to_delete):
        ws.delete_rows(row_idx - offset)
    wb.save(EXCEL_FILE)
    return {"message": "Deal deleted", "trade_id": trade_id}


# --- Prices ---

class PriceEntry(BaseModel):
    product: str
    month: str  # "YYYY-MM"
    price: float
    type: str  # "Settled" or "Forward"
    date: str | None = None  # "YYYY-MM-DD", defaults to today


@app.get("/prices")
def get_prices():
    wb = get_or_create_workbook()
    ws = wb["Prices"]
    prices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        prices.append(dict(zip(PRICE_COLUMNS, row)))
    return prices


@app.post("/prices")
def set_price(entry: PriceEntry):
    wb = get_or_create_workbook()
    ws = wb["Prices"]
    entry_date = entry.date or str(date.today())
    # Update existing match on product + month + type + date, or add new
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_date = row[4].value if len(row) >= 5 else None
        if (row[0].value == entry.product
                and row[1].value == entry.month
                and (len(row) < 4 or row[3].value is None or row[3].value == entry.type)
                and (row_date == entry_date)):
            row[2].value = entry.price
            if len(row) >= 4:
                row[3].value = entry.type
            if len(row) >= 5:
                row[4].value = entry_date
            wb.save(EXCEL_FILE)
            return {"message": "Updated", "product": entry.product, "month": entry.month}
    ws.append([entry.product, entry.month, entry.price, entry.type, entry_date])
    wb.save(EXCEL_FILE)
    return {"message": "Saved", "product": entry.product, "month": entry.month}


@app.post("/prices/bulk")
def bulk_upload_prices(entries: list[PriceEntry]):
    wb = get_or_create_workbook()
    ws = wb["Prices"]
    today = str(date.today())
    # Build index of existing rows for fast lookup
    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        key = (row[0].value, row[1].value,
               row[3].value if len(row) >= 4 else None,
               row[4].value if len(row) >= 5 else None)
        existing[key] = row
    count_updated = 0
    count_added = 0
    for entry in entries:
        entry_date = entry.date or today
        key = (entry.product, entry.month, entry.type, entry_date)
        if key in existing:
            existing[key][2].value = entry.price
            count_updated += 1
        else:
            ws.append([entry.product, entry.month, entry.price, entry.type, entry_date])
            count_added += 1
    wb.save(EXCEL_FILE)
    return {"message": f"{count_added} added, {count_updated} updated"}


# --- ICE PDF Upload ---

# Map ICE commodity codes to our product names
ICE_CODE_TO_PRODUCT = {
    "TFM": "TTF",
}

MONTH_ABBR = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def parse_ice_contract_month(contract_month: str) -> str | None:
    """Convert ICE format like 'Apr26' to 'YYYY-MM' format like '2026-04'."""
    match = re.match(r"([A-Za-z]{3})(\d{2})", contract_month)
    if not match:
        return None
    abbr, year_short = match.group(1), match.group(2)
    month_num = MONTH_ABBR.get(abbr)
    if not month_num:
        return None
    year = f"20{year_short}"
    return f"{year}-{month_num}"


def extract_ice_pdf_date(pdf) -> str | None:
    """Try to extract the trade/report date from an ICE PDF header text."""
    first_page = pdf.pages[0] if pdf.pages else None
    if not first_page:
        return None
    text = first_page.extract_text() or ""
    # Look for patterns like "20-Mar-2026", "March 20, 2026", "2026-03-20", "20 Mar 2026"
    # ICE reports typically use "DD-Mon-YYYY" or "DD Mon YYYY"
    m = re.search(r"(\d{1,2})[-\s](Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-\s](\d{4})", text)
    if m:
        day, mon_abbr, year = m.group(1), m.group(2), m.group(3)
        month_num = MONTH_ABBR.get(mon_abbr)
        if month_num:
            return f"{year}-{month_num}-{int(day):02d}"
    # Try ISO format YYYY-MM-DD
    m = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if m:
        return m.group(1)
    return None


@app.post("/prices/upload-ice-pdf")
async def upload_ice_pdf(file: UploadFile = File(...)):
    """Parse an ICE end-of-day PDF report and save forward prices."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")

    contents = await file.read()
    pdf = pdfplumber.open(io.BytesIO(contents))

    # Extract the report date from the PDF, fall back to today
    pdf_date = extract_ice_pdf_date(pdf) or str(date.today())

    entries = []
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            for row in table:
                if not row or not row[0]:
                    continue
                code = row[0].strip()
                if code not in ICE_CODE_TO_PRODUCT:
                    continue
                product = ICE_CODE_TO_PRODUCT[code]
                # Column 1 = contract month, Column 6 = settle price
                contract_month = (row[1] or "").strip()
                settle_price = (row[6] or "").strip()
                if not contract_month or not settle_price:
                    continue
                month = parse_ice_contract_month(contract_month)
                if not month:
                    continue
                try:
                    price_val = float(settle_price.replace(",", ""))
                except ValueError:
                    continue
                entries.append(PriceEntry(
                    product=product,
                    month=month,
                    price=price_val,
                    type="Forward",
                    date=pdf_date,
                ))

    pdf.close()

    if not entries:
        raise HTTPException(status_code=400, detail="No prices found in PDF")

    # Bulk save using existing logic
    wb = get_or_create_workbook()
    ws = wb["Prices"]
    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        key = (row[0].value, row[1].value,
               row[3].value if len(row) >= 4 else None,
               row[4].value if len(row) >= 5 else None)
        existing[key] = row
    count_updated = 0
    count_added = 0
    for entry in entries:
        key = (entry.product, entry.month, entry.type, entry.date)
        if key in existing:
            existing[key][2].value = entry.price
            count_updated += 1
        else:
            ws.append([entry.product, entry.month, entry.price, entry.type, entry.date])
            count_added += 1
    wb.save(EXCEL_FILE)

    return {
        "message": f"{count_added} added, {count_updated} updated",
        "product": entries[0].product if entries else None,
        "count": len(entries),
        "date": pdf_date,
    }


# --- Settlements ---

@app.get("/settlements/{month}")
def calculate_settlements(month: str):
    """month format: YYYY-MM"""
    year, mon = int(month[:4]), int(month[5:7])
    days_in_month = monthrange(year, mon)[1]
    month_start = date(year, mon, 1)
    month_end = date(year, mon, days_in_month)

    wb = get_or_create_workbook()

    # Load settled prices into a dict (only Type = "Settled")
    # When multiple dates exist, use the latest date's price
    ws_sp = wb["Prices"]
    settled = {}
    settled_dates = {}
    for row in ws_sp.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            price_type = row[3] if len(row) >= 4 else None
            row_date = row[4] if len(row) >= 5 else None
            if price_type == "Settled" or price_type is None:
                key = (row[0], row[1])
                prev_date = settled_dates.get(key)
                if prev_date is None or (row_date or "") >= (prev_date or ""):
                    settled[key] = row[2]
                    settled_dates[key] = row_date

    # Load deals active in this month
    ws_deals = wb["Deals"]
    results = []
    for row in ws_deals.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        deal = dict(zip(COLUMNS, row))
        deal_start = date.fromisoformat(str(deal["Start Date"]))
        deal_end = date.fromisoformat(str(deal["End Date"]))

        # Check if deal is active in this month
        if deal_start > month_end or deal_end < month_start:
            continue

        product = deal["Product"]
        settled_price = settled.get((product, month))
        strike = deal["Price"]
        daily_volume = deal["Volume"]
        direction = deal["Direction"]
        trade_type = deal["Trade Type"]
        monthly_volume = daily_volume * days_in_month

        if settled_price is None:
            pnl = None
        else:
            if trade_type == "Swap":
                if direction == "Buy":
                    pnl = (settled_price - strike) * monthly_volume
                else:
                    pnl = (strike - settled_price) * monthly_volume
            elif trade_type == "Put":
                intrinsic = max(strike - settled_price, 0)
                if direction == "Buy":
                    pnl = intrinsic * monthly_volume
                else:
                    pnl = -intrinsic * monthly_volume
            elif trade_type == "Call":
                intrinsic = max(settled_price - strike, 0)
                if direction == "Buy":
                    pnl = intrinsic * monthly_volume
                else:
                    pnl = -intrinsic * monthly_volume
            else:
                pnl = None

        results.append({
            "Trade ID": deal["Trade ID"],
            "Counterparty": deal["Counterparty"],
            "Direction": direction,
            "Commodity": deal["Commodity"],
            "Product": product,
            "Trade Type": trade_type,
            "Daily Volume": daily_volume,
            "Volume Unit": deal["Volume Unit"],
            "Days": days_in_month,
            "Monthly Volume": monthly_volume,
            "Strike": strike,
            "Settled Price": settled_price,
            "Currency": deal.get("Currency"),
            "P&L": round(pnl, 2) if pnl is not None else None,
        })

    return results
